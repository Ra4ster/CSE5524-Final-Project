import openpyxl as op
import numpy as np
import skimage as sk
from pathlib import Path
from typing import Callable
from lens1 import harris_corners, otsu, ptl_ccl
from scipy.stats import pearsonr

# image_id,model_id,slice_index,z_position_um,r_large_um,porosity_target,porosity_actual,is_heterogeneous,K_x,K_y,K_z,tau_x,tau_y,tau_z,D_x,D_y,D_z,lambda_x,lambda_y,lambda_z,sigma_x,sigma_y,sigma_z,local_slice_porosity,slice_spacing_um,image_path

class ptl_dataset:
    xlsx: Path = Path()

    img_ids: list[str] = []
    model_ids: list[str] = []
    slice_idx: list[int] = []
    z_pos_um: list[int] = []
    r_lg_um: list[int] = []
    porosity_target: list[float] = []
    porosity_actual: list[float] = []
    is_heterogenous: list[bool] = []
    K_x: list[float] = []
    K_y: list[float] = []
    K_z: list[float] = []
    tau_x: list[float] = []
    tau_y: list[float] = []
    tau_z: list[float] = []
    D_x: list[float] = []
    D_y: list[float] = []
    D_z: list[float] = []
    lambda_x: list[float] = []
    lambda_y: list[float] = []
    lambda_z: list[float] = []
    sigma_x: list[float] = []
    sigma_y: list[float] = []
    sigma_z: list[float] = []
    local_slice_porosity: list[float] = []
    slice_spacing_um: list[float] = []
    img_path: list[str] = []
    images: list[np.ndarray] = []

    workbook: op.Workbook = op.Workbook()

    @staticmethod
    def _to_str(val):
        try:
            return str(val) if val is not None else ""
        except:
            return ""

    @staticmethod
    def _to_int(val):
        try:
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return int(val)
            return int(float(str(val)))
        except:
            return 0

    @staticmethod
    def _to_float(val):
        try:
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            return float(str(val))
        except:
            return 0.0

    @staticmethod
    def _to_bool(val):
        try:
            if val is None:
                return False
            if isinstance(val, bool):
                return val
            if isinstance(val, (int, float)):
                return bool(val)
            return str(val).lower() in ("true", "1", "yes")
        except:
            return False

    def __init__(self, xlsx: str, rows: int | None = None):
        self.workbook = op.load_workbook(filename=xlsx, read_only=True)
        self.xlsx = Path(xlsx)
        assert self.workbook.active, "Workbook must be active."
        assert rows == None or rows > 0, "Rows must be a natural number."
        sheet = self.workbook.active
        if rows == None:
            rows: int = sheet.max_row
            print(f"[DEBUG] There are {rows} rows in {self.xlsx.name}")
        
        # Populate lists from workbook (skip header row)
        for row in sheet.iter_rows(min_row=2, values_only=True, max_row=rows):
            self.img_ids.append(self._to_str(row[0]))
            self.model_ids.append(self._to_str(row[1]))
            self.slice_idx.append(self._to_int(row[2]))
            self.z_pos_um.append(self._to_int(row[3]))
            self.r_lg_um.append(self._to_int(row[4]))
            self.porosity_target.append(self._to_float(row[5]))
            self.porosity_actual.append(self._to_float(row[6]))
            self.is_heterogenous.append(self._to_bool(row[7]))
            self.K_x.append(self._to_float(row[8]))
            self.K_y.append(self._to_float(row[9]))
            self.K_z.append(self._to_float(row[10]))
            self.tau_x.append(self._to_float(row[11]))
            self.tau_y.append(self._to_float(row[12]))
            self.tau_z.append(self._to_float(row[13]))
            self.D_x.append(self._to_float(row[14]))
            self.D_y.append(self._to_float(row[15]))
            self.D_z.append(self._to_float(row[16]))
            self.lambda_x.append(self._to_float(row[17]))
            self.lambda_y.append(self._to_float(row[18]))
            self.lambda_z.append(self._to_float(row[19]))
            self.sigma_x.append(self._to_float(row[20]))
            self.sigma_y.append(self._to_float(row[21]))
            self.sigma_z.append(self._to_float(row[22]))
            self.local_slice_porosity.append(self._to_float(row[23]))
            self.slice_spacing_um.append(self._to_float(row[24]))
            self.img_path.append(self._to_str(row[25]))

        for fp in self.img_path:
            self.images.append(
                sk.io.imread(
                    Path("../" + fp)
                    )
                )

    def cmp_porosity_fn(self, fn: Callable[[np.ndarray], float]) -> list[float]:
        diffs: list[float] = []
        for i,img in enumerate(self.images):
            attm: float = fn(img)
            attm_porosity: float = np.sum(img > attm) / img.size
            diffs.append(attm_porosity - self.porosity_actual[i])
        return diffs
    
    def cmp_porosities(self, result: list[float]) -> list[float]:
        return [self.porosity_actual[i] - result[i] for i in range(len(result))]


def main():
    testdata = ptl_dataset(xlsx="../ptl_cv_dataset/PTL_CV_Dataset.xlsx")

    porosities: list[float] = []
    diffs: list[float] = []
    elongations: list[float] = []
    junction_densities: list[float] = []

    print("Processing dataset and extracting features...")
    for i, img in enumerate(testdata.images):
        t, _ = otsu(img)
        
        porosity = np.sum(img > t) / img.size
        porosities.append(porosity)
        
        a = testdata.local_slice_porosity[i]
        diffs.append(a - porosity) # Ground truth
        
        mask = img < t # This may take a while!
        ccl = ptl_ccl(mask, c=2, min_pixels=20)
        elongations.append(ccl.average_elongation())

        # Junction Density
        hc = harris_corners(img.astype(float) / 255.)
        density = len(hc) / img.size
        junction_densities.append(density)

    # MSE
    mse = np.sum([diff**2 for diff in diffs]) / len(diffs)
    print(f"[RESULT] Otsu's True 2D MSE: {mse}")

    # Correlations
    corr_tau_z, p_tau = pearsonr(elongations, testdata.tau_z)
    corr_radius, p_rad = pearsonr(junction_densities, testdata.r_lg_um)

    print(f"[RESULT] Elongation vs. Tortuosity (Z) Correlation: {corr_tau_z:.4f}")
    print(f"[RESULT] Junction Density vs. Fiber Radius Correlation: {corr_radius:.4f}")

    print("\nExporting results to Excel...")
    out_wb = op.Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Lens 1 Results"

    headers = [
        "Image_ID", 
        "True_2D_Porosity", 
        "Otsu_Porosity", 
        "Porosity_Error", 
        "Average_Elongation", 
        "Junction_Density"
    ]
    out_sheet.append(headers)

    for i in range(len(testdata.images)):
        row_data = [
            testdata.img_ids[i],
            testdata.local_slice_porosity[i],
            porosities[i],
            diffs[i],
            elongations[i],
            junction_densities[i]
        ]
        out_sheet.append(row_data)

    output_filename = "Lens1_Results.xlsx"
    out_wb.save(output_filename)
    print(f"[RESULT] Successfully saved to {output_filename}")


if __name__ == "__main__":
    main()